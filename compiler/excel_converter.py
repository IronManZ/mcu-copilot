#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel格式转换工具
将Excel工作簿中的汇编代码转换为标准文本格式
"""

import sys
import json
from pathlib import Path
from typing import Dict, List, Optional

try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

class ExcelToTextConverter:
    """Excel到文本格式转换器"""
    
    def __init__(self):
        self.variables = {}
        self.instructions = []
        self.debug_info = []
    
    def convert_file(self, excel_file: str, output_file: str = None) -> str:
        """转换Excel文件到文本格式"""
        if not EXCEL_SUPPORT:
            raise ImportError("需要安装openpyxl库: pip install openpyxl")
        
        try:
            # 读取Excel文件
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            
            # 分析工作表结构
            self._analyze_workbook_structure(wb)
            
            # 解析Code工作表
            if 'Code' in wb.sheetnames:
                self._parse_code_sheet(wb['Code'])
            else:
                raise ValueError("未找到Code工作表")
            
            # 生成文本格式
            text_content = self._generate_text_format()
            
            # 保存文件
            if output_file:
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(text_content)
                print(f"✓ 转换完成，已保存到: {output_file}")
            
            return text_content
            
        except Exception as e:
            raise Exception(f"转换失败: {str(e)}")
    
    def _analyze_workbook_structure(self, wb):
        """分析工作簿结构"""
        self.debug_info.append(f"工作表列表: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            self.debug_info.append(f"{sheet_name}: {max_row}行 x {max_col}列")
    
    def _parse_code_sheet(self, sheet):
        """解析Code工作表"""
        print("正在解析Code工作表...")
        
        in_data_section = False
        in_code_section = False
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if not row or all(cell is None or cell == '' for cell in row):
                continue
            
            # 转换为字符串列表，处理None值
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append('')
                elif isinstance(cell, (int, float)):
                    row_data.append(str(int(cell)) if isinstance(cell, float) and cell.is_integer() else str(cell))
                else:
                    row_data.append(str(cell))
            
            # 调试输出前几行
            if row_idx <= 20:
                print(f"行{row_idx}: {row_data[:6]}")  # 只显示前6列
            
            # 检查段标识符
            first_col = row_data[0].strip() if row_data[0] else ''
            
            if first_col == 'DATA':
                in_data_section = True
                in_code_section = False
                print("  -> 进入DATA段")
                continue
            elif first_col == 'ENDDATA':
                in_data_section = False
                print("  -> 离开DATA段")
                continue
            elif first_col == 'CODE':
                in_code_section = True
                in_data_section = False
                print("  -> 进入CODE段")
                continue
            elif first_col == 'ENDCODE':
                in_code_section = False
                print("  -> 离开CODE段")
                continue
            
            # 解析DATA段
            if in_data_section and len(row_data) > 2:
                var_name = row_data[1].strip() if row_data[1] else ''
                address_str = row_data[2].strip() if row_data[2] else ''
                
                if var_name and address_str:
                    try:
                        # 处理可能的浮点数地址
                        address = int(float(address_str))
                        self.variables[var_name] = address
                        print(f"  变量: {var_name} = {address}")
                    except (ValueError, TypeError):
                        print(f"  警告: 无法解析变量地址 {var_name}: {address_str}")
            
            # 解析CODE段
            elif in_code_section and len(row_data) > 1:
                # Excel格式: 列A(空), 列B(标号), 列C(指令), 列D(操作数)
                label = row_data[1].strip() if len(row_data) > 1 and row_data[1] else None
                mnemonic = row_data[2].strip() if len(row_data) > 2 and row_data[2] else ''
                operand = row_data[3].strip() if len(row_data) > 3 and row_data[3] else ''
                
                # 清理标号格式（移除可能的冒号）
                if label and label.endswith(':'):
                    label = label[:-1]
                
                # 过滤掉空行和无效行
                if mnemonic or label:
                    instruction = {
                        'row': row_idx,
                        'label': label if label else None,
                        'mnemonic': mnemonic,
                        'operand': operand
                    }
                    self.instructions.append(instruction)
                    
                    label_str = f"{label}:" if label else ""
                    # 特殊处理DB指令的显示
                    if mnemonic == 'DB':
                        print(f"  数据: {label_str:12s} {mnemonic:8s} {operand}")
                    else:
                        print(f"  指令: {label_str:12s} {mnemonic:8s} {operand}")
        
        print(f"解析完成: {len(self.variables)}个变量, {len(self.instructions)}条指令")
    
    def _generate_text_format(self) -> str:
        """生成标准文本格式"""
        lines = []
        
        # 生成DATA段
        if self.variables:
            lines.append('DATA')
            # 按地址排序变量
            sorted_vars = sorted(self.variables.items(), key=lambda x: x[1])
            for var_name, address in sorted_vars:
                lines.append(f'    {var_name:<12} {address}')
            lines.append('ENDDATA')
            lines.append('')
        
        # 生成CODE段
        if self.instructions:
            lines.append('CODE')
            
            for i, inst in enumerate(self.instructions):
                label = inst['label']
                mnemonic = inst['mnemonic']
                operand = inst['operand']
                
                # 处理标号和指令的格式
                if label and mnemonic:
                    # 标号和指令都存在
                    # 先输出标号，然后下一行缩进输出指令
                    lines.append(f'{label}:')
                    if operand:
                        lines.append(f'    {mnemonic} {operand}')
                    else:
                        lines.append(f'    {mnemonic}')
                elif label and not mnemonic:
                    # 只有标号，标号单独一行（等待下一条指令）
                    lines.append(f'{label}:')
                elif mnemonic:
                    # 只有指令，使用缩进
                    if operand:
                        lines.append(f'    {mnemonic} {operand}')
                    else:
                        lines.append(f'    {mnemonic}')
                
                # 在标号后添加空行以提高可读性
                if label and i < len(self.instructions) - 1:
                    # 检查下一条指令是否也有标号
                    next_inst = self.instructions[i + 1]
                    if next_inst.get('label'):
                        lines.append('')  # 在标号间添加空行
            
            lines.append('ENDCODE')
        
        return '\n'.join(lines)
    
    def get_debug_info(self) -> List[str]:
        """获取调试信息"""
        return self.debug_info

class TextToExcelConverter:
    """文本到Excel格式转换器"""
    
    def __init__(self):
        self.variables = {}
        self.instructions = []
    
    def convert_file(self, text_file: str, output_file: str = None) -> str:
        """转换文本文件到Excel格式"""
        if not EXCEL_SUPPORT:
            raise ImportError("需要安装openpyxl库: pip install openpyxl")
        
        try:
            # 读取文本文件
            with open(text_file, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 解析文本格式
            self._parse_text_format(content)
            
            # 生成Excel文件
            excel_file = output_file or text_file.replace('.asm', '.xlsx')
            self._generate_excel_file(excel_file)
            
            print(f"✓ 转换完成，已保存到: {excel_file}")
            return excel_file
            
        except Exception as e:
            raise Exception(f"转换失败: {str(e)}")
    
    def _parse_text_format(self, content: str):
        """解析文本格式"""
        lines = content.split('\n')
        in_data_section = False
        in_code_section = False
        
        for line_no, line in enumerate(lines, 1):
            line = line.strip()
            
            if not line or line.startswith(';') or line.startswith("'"):
                continue
            
            if line == 'DATA':
                in_data_section = True
                in_code_section = False
                continue
            elif line == 'ENDDATA':
                in_data_section = False
                continue
            elif line == 'CODE':
                in_code_section = True
                in_data_section = False
                continue
            elif line == 'ENDCODE':
                in_code_section = False
                continue
            
            if in_data_section:
                parts = line.split()
                if len(parts) >= 2:
                    var_name = parts[0]
                    try:
                        address = int(parts[1])
                        self.variables[var_name] = address
                    except ValueError:
                        pass
            
            elif in_code_section:
                label = None
                if ':' in line:
                    parts = line.split(':', 1)
                    label = parts[0].strip()
                    line = parts[1].strip() if len(parts) > 1 else ''
                
                # 处理缩进的指令（移除前导空格来解析指令）
                line = line.lstrip()
                
                if line:
                    parts = line.split()
                    mnemonic = parts[0] if parts else ''
                    operand = ' '.join(parts[1:]) if len(parts) > 1 else ''
                    
                    self.instructions.append({
                        'line_no': line_no,
                        'label': label,
                        'mnemonic': mnemonic,
                        'operand': operand
                    })
                elif label:
                    # 只有标号的行
                    self.instructions.append({
                        'line_no': line_no,
                        'label': label,
                        'mnemonic': '',
                        'operand': ''
                    })
    
    def _generate_excel_file(self, filename: str):
        """生成Excel文件"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Code"
        
        # 设置标题行
        ws.append(['', '变量名/标号', '地址/指令', '操作数', '', '', '', '', '', '备注'])
        
        # 写入DATA段
        if self.variables:
            ws.append(['DATA', '变量名', '地址'])
            for var_name, address in sorted(self.variables.items(), key=lambda x: x[1]):
                ws.append(['', var_name, address])
            ws.append(['ENDDATA'])
            ws.append([''])
        
        # 写入CODE段
        if self.instructions:
            ws.append(['CODE', '标号', '指令', '操作数'])
            for inst in self.instructions:
                ws.append([
                    '',
                    inst['label'] or '',
                    inst['mnemonic'],
                    inst['operand']
                ])
            ws.append(['ENDCODE'])
        
        wb.save(filename)

def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Excel格式转换工具')
    parser.add_argument('input', help='输入文件路径')
    parser.add_argument('-o', '--output', help='输出文件路径')
    parser.add_argument('-m', '--mode', choices=['excel-to-text', 'text-to-excel'], 
                       required=True, help='转换模式')
    parser.add_argument('-d', '--debug', action='store_true', help='显示调试信息')
    
    args = parser.parse_args()
    
    try:
        if args.mode == 'excel-to-text':
            converter = ExcelToTextConverter()
            result = converter.convert_file(args.input, args.output)
            
            if args.debug:
                print("\n=== 调试信息 ===")
                for info in converter.get_debug_info():
                    print(info)
                
                print(f"\n变量列表 ({len(converter.variables)} 个):")
                for name, addr in converter.variables.items():
                    print(f"  {name}: {addr}")
                
                print(f"\n指令列表 ({len(converter.instructions)} 条):")
                for i, inst in enumerate(converter.instructions[:10]):  # 显示前10条
                    print(f"  {i}: {inst}")
                if len(converter.instructions) > 10:
                    print(f"  ... 还有 {len(converter.instructions) - 10} 条")
        
        elif args.mode == 'text-to-excel':
            converter = TextToExcelConverter()
            converter.convert_file(args.input, args.output)
        
    except Exception as e:
        print(f"错误: {str(e)}")
        sys.exit(1)

def example_usage():
    """使用示例"""
    print("=== Excel格式转换工具示例 ===")
    
    if not EXCEL_SUPPORT:
        print("需要安装openpyxl库: pip install openpyxl")
        return
    
    # 示例：创建一个文本文件然后转换为Excel
    sample_text = """DATA
    counter    0
    result     1
    temp       2
ENDDATA

CODE
start:
    LDINS 10
    ST counter

loop:
    LD counter
    DEC
    ST counter
    JZ end
    JUMP loop

end:
    LD result
    NOP
ENDCODE
"""
    
    # 保存示例文本文件
    with open('sample.asm', 'w', encoding='utf-8') as f:
        f.write(sample_text)
    
    try:
        # 文本转Excel
        text_converter = TextToExcelConverter()
        excel_file = text_converter.convert_file('sample.asm', 'sample.xlsx')
        
        # Excel转文本 (验证往返转换)
        excel_converter = ExcelToTextConverter()
        converted_text = excel_converter.convert_file('sample.xlsx', 'sample_converted.asm')
        
        print("\n=== 转换结果对比 ===")
        print("原始文本:")
        print(sample_text[:200] + "...")
        print("\n转换后文本:")
        print(converted_text[:200] + "...")
        
        print(f"\n变量数量: {len(excel_converter.variables)}")
        print(f"指令数量: {len(excel_converter.instructions)}")
        
    except Exception as e:
        print(f"示例运行失败: {str(e)}")

if __name__ == '__main__':
    if len(sys.argv) > 1:
        main()
    else:
        example_usage()